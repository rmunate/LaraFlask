import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from lib.clarity.file import File

class Excel:
    """
    Esta clase permite generar archivos de Excel de forma rapida y eficiente,
    suministra una interfaz elegante para determinar las acciones de insercion de datos, definicion 
    de estilos y cabeceras.
    La salida siempre será un archivo fisico en el servidor.
    """

    def __init__(self) -> None:
        """Inicializa la clase de Excel"""
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def setBackgroundColorHeader(self, color:str):
        """
        Define el color de fondo que se usará en los encabezados.

        Args:
            color (str): El color en formato hexadecimal que se usará como fondo en los encabezados del excel.

        Returns:
            Excel: La instancia actual de la clase Excel.
        """
        self.headers_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        return self

    def setTextBoldHeader(self, bold:bool = False):
        """
        Determina si el texto de los encabezados estará en negrilla.

        Args:
            bold (bool, optional): Indica si el texto de los encabezados estará en negrilla. Por defecto es False.

        Returns:
            Excel: La instancia actual de la clase Excel.
        """
        self.headers_style = Font(bold=bold)
        return self

    def setTextColorHeader(self, color:str):
        """
        Determina si el texto de los encabezados estará en negrilla.

        Args:
            bold (bool, optional): Indica si el texto de los encabezados estará en negrilla. Por defecto es False.

        Returns:
            Excel: La instancia actual de la clase Excel.
        """
        self.headers_style.color = color
        return self

    def setSheetName(self, sheet_name:str):
        """
        Define el color del texto de los encabezados.

        Args:
            color (str): El color en formato hexadecimal del texto de los encabezados.

        Returns:
            Excel: La instancia actual de la clase Excel.
        """
        self.sheet.title = sheet_name
        return self

    def headers(self, headers):
        """
        Define los encabezados del Excel.

        Args:
            headers (list): Lista de encabezados a escribir.

        Returns:
            Excel: La instancia actual de la clase Excel.
        """
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = self.sheet[f'{col_letter}1']
            cell.value = header
            cell.font = self.headers_style
            cell.fill = self.headers_fill
        return self

    def rows(self, data):
        """
        Asigna los valores de las filas.

        Args:
            data (list): Lista de datos a escribir en las filas.

        Returns:
            Excel: La instancia actual de la clase Excel.
        """
        for row in data:
            self.sheet.append(row)
        return self

    def save(self, filename:str):
        """
        Crea el archivo de Excel y devuelve la ruta del archivo.

        Args:
            filename (str): El nombre del archivo de Excel.

        Returns:
            str: La ruta del archivo de Excel creado.
        """
        path = File.path(name=filename)
        self.workbook.save(path)
        return path
